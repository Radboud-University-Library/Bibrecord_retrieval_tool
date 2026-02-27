"""
export.py

This module handles the export of MARCXML and JSON data to Excel files.
It includes functions for parsing MARCXML data, exporting data to Excel,
merging Excel files, and saving XML files to a ZIP archive.

Functions:
- parse_marcxml(xml_data)
- export_xml_data_to_excel(xml_directory, final_filename, progress_bar)
- export_json_data_to_excel(json_directory, final_filename, progress_bar)
- merge_excel_files(xml_filename, json_filename, merged_filename)
- save_all_xml_to_zip(xml_directory, zip_file_name)
"""

import streamlit as st
import os
import xml.etree.ElementTree as ET
import pandas as pd
import zipfile
import json
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def _load_single_json_object(filepath):
    """Load a JSON file, tolerating trailing garbage after the first object."""
    with open(filepath, 'r', encoding='utf-8') as file:
        text = file.read()
    try:
        return json.loads(text), None
    except json.JSONDecodeError:
        # Fallback: parse the first object and ignore anything after it
        decoder = json.JSONDecoder()
        stripped = text.lstrip()
        obj, idx = decoder.raw_decode(stripped)
        trailing = stripped[idx:].strip()
        return obj, trailing if trailing else None


def parse_marcxml(xml_data):
    """Parse MARCXML data and return a list of dictionaries."""
    ns = {'marc': 'http://www.loc.gov/MARC21/slim'}
    root = ET.fromstring(xml_data)
    all_records = []

    for record in [root]:
        record_dict = {}
        for cf in record.findall('marc:controlfield', ns):
            tag = cf.get('tag')
            record_dict[tag] = cf.text.strip() if cf.text else ''

        for df in record.findall('marc:datafield', ns):
            tag = df.get('tag')
            for sf in df.findall('marc:subfield', ns):
                code = sf.get('code')
                value = sf.text.strip() if sf.text else ''
                key = f"{tag}_{code}"
                if key in record_dict:
                    if isinstance(record_dict[key], list):
                        record_dict[key].append(value)
                    else:
                        record_dict[key] = [record_dict[key], value]
                else:
                    record_dict[key] = value

        all_records.append(record_dict)

    return all_records


def export_xml_data_to_excel(xml_directory, final_filename, progress_bar):

    if not os.path.exists(xml_directory):
        st.error("No XML files found to export.")
        return

    xml_files = [f for f in os.listdir(xml_directory) if f.endswith('.xml')]
    if not xml_files:
        st.error("No XML files found to export.")
        return

    total_files = len(xml_files)
    chunk_size = 5000

    # First pass: determine the complete set of columns across all files
    all_columns = set()
    for filename in xml_files:
        filepath = os.path.join(xml_directory, filename)
        with open(filepath, 'r', encoding='utf-8') as file:
            xml_data = file.read()
            parsed_data = parse_marcxml(xml_data)
            for record in parsed_data:
                all_columns.update(record.keys())
    # Fix the order of columns (for example, sorted alphabetically)
    all_columns = sorted(all_columns)

    # Second pass: process the XML files in chunks with a consistent column order
    for i in range(0, total_files, chunk_size):
        chunk_data = []
        chunk = xml_files[i:i + chunk_size]
        for filename in chunk:
            filepath = os.path.join(xml_directory, filename)
            with open(filepath, 'r', encoding='utf-8') as file:
                xml_data = file.read()
                parsed_data = parse_marcxml(xml_data)
                chunk_data.extend(parsed_data)

        # Create DataFrame and reindex using the pre-determined column order
        df = pd.DataFrame(chunk_data)
        df = df.reindex(columns=all_columns)  # This ensures consistent alignment

        # Write the data to the final Excel file
        if i == 0:  # For the first chunk, create a new file with header
            df.to_excel(final_filename, index=False, engine='openpyxl')
        else:  # For subsequent chunks, append to the existing sheet
            with pd.ExcelWriter(final_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                startrow = writer.book['Sheet1'].max_row
                df.to_excel(writer, index=False, header=False, startrow=startrow)

        progress_bar.progress((i + len(chunk)) / total_files, text=f"XML {min(i+len(chunk), total_files)}/{total_files}")

    st.success(f"XML data has been exported to {final_filename}")


def export_json_data_to_excel(json_directory, final_filename, progress_bar):
    if not os.path.exists(json_directory):
        st.error("No JSON files found to export.")
        return

    json_files = [f for f in os.listdir(json_directory) if f.endswith('.json')]
    if not json_files:
        st.error("No JSON files found to export.")
        return

    total_files = len(json_files)
    all_data = []

    # Process each JSON file
    for i, filename in enumerate(json_files):
        filepath = os.path.join(json_directory, filename)
        try:
            json_data, trailing = _load_single_json_object(filepath)
        except json.JSONDecodeError as e:
            st.warning(f"Skipping invalid JSON file: {filename} ({e})")
            continue

        if trailing:
            st.warning(f"Ignoring trailing data after JSON object in {filename}")

        if isinstance(json_data, dict) and "holdings" in json_data:
            base_record = {"ocn": json_data.get("ocn")}
            holdings = json_data["holdings"]

            # Flatten ONLY totalHoldingCount per institution
            for holding in holdings:
                symbol = holding.get("institutionSymbol")
                if symbol is None:
                    continue
                total_count = holding.get("totalHoldingCount", 0)
                column_name = f"totalHoldingCount_{symbol}"
                base_record[column_name] = total_count

            all_data.append(base_record)

        # Update progress bar
        progress_bar.progress((i + 1) / total_files, text=f"JSON {i+1}/{total_files}")

    # Write the flattened JSON data to the final Excel file
    df = pd.DataFrame(all_data)
    df.to_excel(final_filename, index=False, engine='openpyxl')

    st.success(f"JSON data has been exported to {final_filename}")


def _normalize_ocn(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    match = re.search(r'(\d+)', text)
    if not match:
        return ""
    try:
        return str(int(match.group(1)))
    except Exception:
        return match.group(1)


def _merge_excel_files_streaming(xml_filename, json_filename, merged_filename):
    # Build a holdings map from JSON export (ocn -> list of totalHoldingCount_* values)
    wb_json = load_workbook(json_filename, read_only=True, data_only=True)
    ws_json = wb_json.active
    json_header = next(ws_json.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not json_header:
        wb_json.close()
        st.error("JSON export is empty. Cannot merge.")
        return merged_filename

    try:
        ocn_idx = list(json_header).index("ocn")
    except ValueError:
        wb_json.close()
        st.error("JSON export does not contain an 'ocn' column.")
        return merged_filename

    json_cols = [c for c in json_header if isinstance(c, str) and c.startswith("totalHoldingCount_")]
    json_col_indices = [list(json_header).index(c) for c in json_cols]

    holdings_map = {}
    for row in ws_json.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        ocn = _normalize_ocn(row[ocn_idx] if ocn_idx < len(row) else None)
        if not ocn:
            continue
        values = [row[i] if i < len(row) else None for i in json_col_indices]
        holdings_map[ocn] = values
    wb_json.close()

    # Stream XML rows and write merged output to a new workbook
    wb_xml = load_workbook(xml_filename, read_only=True, data_only=True)
    ws_xml = wb_xml.active
    xml_header = next(ws_xml.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not xml_header:
        wb_xml.close()
        st.error("XML export is empty. Cannot merge.")
        return merged_filename

    try:
        xml_ocn_idx = list(xml_header).index("001")
    except ValueError:
        wb_xml.close()
        st.error("XML export does not contain a '001' column.")
        return merged_filename

    include_sum = len(json_cols) > 0
    sum_col_name = "totalHoldingCount_SUM"
    out_header = list(xml_header)
    if include_sum:
        # Add a matched OCN column between XML and JSON for easy validation
        out_header.append("ocn")
        out_header.append(sum_col_name)
        out_header.extend(json_cols)

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title="Sheet1")
    ws_out.append(out_header)

    # Precompute formula column letters if needed
    if include_sum:
        ocn_col_idx = len(xml_header) + 1
        sum_col_idx = ocn_col_idx + 1
        first_json_idx = sum_col_idx + 1
        last_json_idx = first_json_idx + len(json_cols) - 1
        first_json_letter = get_column_letter(first_json_idx)
        last_json_letter = get_column_letter(last_json_idx)

    out_row_num = 2  # header is row 1
    for row in ws_xml.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        xml_values = list(row)
        ocn = _normalize_ocn(xml_values[xml_ocn_idx] if xml_ocn_idx < len(xml_values) else None)
        has_match = ocn in holdings_map
        json_values = holdings_map.get(ocn, [None] * len(json_cols)) if include_sum else []

        if include_sum:
            # Use a formula so totals stay dynamic in Excel
            formula = f"=SUM({first_json_letter}{out_row_num}:{last_json_letter}{out_row_num})"
            out_row = xml_values + [ocn if has_match else ""] + [formula] + json_values
        else:
            out_row = xml_values

        ws_out.append(out_row)
        out_row_num += 1

    wb_xml.close()
    wb_out.save(merged_filename)
    st.success(f"XML and JSON data have been merged and exported to {merged_filename}")
    return merged_filename


def merge_excel_files(xml_filename, json_filename, merged_filename):
    """Merge XML + JSON exports using the streaming implementation."""
    return _merge_excel_files_streaming(xml_filename, json_filename, merged_filename)


def save_all_xml_to_zip(xml_directory, zip_file_name):
    """Save all XML files from the specified directory to a ZIP file to Desktop."""
    home_directory = os.path.expanduser('~')
    desktop_directory = os.path.join(home_directory, 'Desktop')

    if not os.path.isdir(desktop_directory):
        st.error("Unable to find the desktop directory.")
        return

    save_location = os.path.join(desktop_directory, zip_file_name)
    with zipfile.ZipFile(save_location, 'w') as zipf:
        for root, _, files in os.walk(xml_directory):
            for file in files:
                if file.endswith('.xml'):
                    zipf.write(os.path.join(root, file), arcname=file)

    st.success(f"All XML files have been saved to {save_location} successfully.")
